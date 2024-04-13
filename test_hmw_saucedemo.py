from selenium import webdriver
import pytest
import time
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as ec #hangi şarta baglı olarak bekeleyecegini
from pathlib import Path
from datetime import date
import openpyxl
from selenium.webdriver.support.wait import WebDriverWait #ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
from selenium.webdriver.common.action_chains import ActionChains 
from constants.globalConstants import *

class Test_Saucedemocom:
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.maximize_window
        self.driver.get(BASE_URL)
    
    def teardown_method(self):
        self.driver.quit()
        
    def waitForElementVisiable(self,locator,timeout = 5):
        WebDriverWait(self.driver, timeout).until(ec.visibility_of_element_located(locator))   
        
    
    def getData():
        return [("1","1"),("abc","123"),("deneme","secret_sauce")]
    
    def readInvalidDataFromExcel():
        excelFile = openpyxl.load_workbook("TOBETO_Selenium/data/D_invalidLogin.xlsx")
        sheet = excelFile["Sheet1"]
        rows = sheet.max_row #kaçıncı satıra kadar benim verim var
        data = []
        for i in range(2, rows+1):
            username = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            data.append((username,password))
        return data 
 
    @pytest.mark.parametrize("username, password", readInvalidDataFromExcel())  
    def test_invalid_login(self, username, password):
        userNameInput = self.driver.find_element(By.ID, USER_NAME_ID)
        passwordInput = self.driver.find_element(By.ID, PASSWORD_ID)
        userNameInput.send_keys(username)
        sleep(5)
        passwordInput.send_keys(password)
        loginButton = self.driver.find_element(By.ID, "login-button")
        loginButton.click()
        errorMessage = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.XPATH, "//*[@id='login_button_container']/div/form/div[3]")))
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"
        
    
        
    def test_basket_product_number(self):   
        userNameInput = self.driver.find_element(By.ID, USER_NAME_ID)
        userNameInput.send_keys("standard_user")
        sleep(2)
        passwordInput = self.driver.find_element(By.ID, PASSWORD_ID)
        passwordInput.send_keys("secret_sauce")
        sleep(2)
        loginButton = self.driver.find_element(By.ID, LOGIN_BUTTON_ID)
        loginButton.click() 
        addButton = self.driver.find_element(By.ID, "add-to-cart-sauce-labs-fleece-jacket")
        addButton.click()
        sleep(5)
        current_basket_locater = self.driver.find_element(By.CLASS_NAME, "shopping_cart_link")
        expected_basket_locetor = self.driver.find_element(By.CLASS_NAME, "shopping_cart_badge")
        assert expected_basket_locetor.is_displayed
     
            
    def test_add_basket(self):
        userNameInput = self.driver.find_element(By.ID, USER_NAME_ID)
        self.waitForElementVisiable(((By.ID,"user-name")),10)
        userNameInput.send_keys("standard_user")
       
        self.waitForElementVisiable(((By.ID,"password")),10)
        passwordInput = self.driver.find_element(By.ID, PASSWORD_ID)
        passwordInput.send_keys("secret_sauce")
      
        loginButton = self.driver.find_element(By.ID, LOGIN_BUTTON_ID)
        loginButton.click() 
        addButton = self.driver.find_element(By.ID, "add-to-cart-sauce-labs-fleece-jacket")
        addButton.click()
        expected_button = self.driver.find_element(By.ID, "remove-sauce-labs-fleece-jacket")
        assert expected_button.is_displayed()   
          
    def test_test_empty_username_password_login(self):
       # loginButton = self.driver.find_element(By.ID, "login-button")
        loginButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID, "login-button")))
        loginButton.click()
        sleep(2)
        #errorMassage = self.driver.find_element(By.ID, "login_button_container")
        errorMassage = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID, "login_button_container")))
        test_result = "Epic sadface: Username is required"
        assert test_result == errorMassage.text
        sleep(2)
      
    def test_emptyPassword_login_test(self): 
      userNameInput = self.driver.find_element(By.ID, USER_NAME_ID)
      userNameInput.send_keys("standard_user")
     
      loginButton = self.driver.find_element(By.ID, LOGIN_BUTTON_ID)
      loginButton.click()
      errorMassage = self.driver.find_element(By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")
      expected_error_message = "Epic sadface: Password is required"
      assert errorMassage.text == expected_error_message
  
   
    def test_exist_userName_password_login(self):
        userNameInput = self.driver.find_element(By.ID, USER_NAME_ID)
        userNameInput.send_keys("locked_out_user")
        passwordInput = self.driver.find_element(By.ID, PASSWORD_ID)
        passwordInput.send_keys("secret_sauce")
        loginButton = self.driver.find_element(By.ID, LOGIN_BUTTON_ID)
        loginButton.click()
        errorMassage = self.driver.find_element(By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")
        testResult = "Epic sadface: Sorry, this user has been locked out."
        assert errorMassage.text == testResult
        
    def test_successful_login_and_product_display(self):
         userNameInput = self.driver.find_element(By.ID, USER_NAME_ID)
         userNameInput.send_keys("standard_user")
         sleep(2)
         passwordInput = self.driver.find_element(By.ID, PASSWORD_ID)
         passwordInput.send_keys("secret_sauce")
         sleep(2)
         #loginButton = self.driver.find_element(By.ID, login_button_id)
         loginButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,LOGIN_BUTTON_ID)))
         loginButton.click()
    
         expectedUrl = "https://www.saucedemo.com/inventory.html"
         
         current_url = self.driver.current_url
         assert current_url == expectedUrl
          
         products = self.driver.find_elements(By.CLASS_NAME, "inventory_item_description")
         #Bu kod, belirli bir CSS sınıfına sahip bir öğenin görünür olmasını bekler ve bu öğeyi döndürür. 
         # Yani products değişkeni, yalnızca bir öğe içerecek, çünkü visibility_of_element_located metodu yalnızca bir öğe döndürür.
         #Eğer sayfada birden fazla öğe olması ve hepsini bulmak istiyorsanız, ilk kod bloğunu kullanmalısınız.
         # Bunu değil : products = WebDriverWait(self.driver,4).until(ec.visibility_of_element_located((By.CLASS_NAME, "inventory_item_description")))
         sleep(3)
         currentProductsNumber = len(products)
         assert currentProductsNumber == 6
         
         

   
    



